"""
NotingHill — services/extractors/xlsx_extractor.py
"""
from pathlib import Path
from .base import BaseExtractor, ExtractionResult


class XlsxExtractor(BaseExtractor):
    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in {".xlsx", ".xls", ".xlsm"}

    def extract(self, path: Path) -> ExtractionResult:
        if self._too_large(path):
            return ExtractionResult(warnings=["File too large"])
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            texts = []
            sheet_names = wb.sheetnames
            for sname in sheet_names[:10]:
                ws = wb[sname]
                texts.append(f"[Sheet: {sname}]")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(c) for c in row if c is not None)
                    if row_text.strip():
                        texts.append(row_text)
                    row_count += 1
                    if row_count > 500:
                        break
            wb.close()
            text = "\n".join(texts)
            r = ExtractionResult(
                text=text,
                metadata={"sheets": sheet_names}
            )
            r.make_preview()
            return r
        except ImportError:
            return ExtractionResult(warnings=["openpyxl not installed"])
        except Exception as e:
            return ExtractionResult(warnings=[str(e)])
